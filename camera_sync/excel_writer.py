# -*- coding: utf-8 -*-
"""Excel 报告写出：与原 CS8K_n_Timecode.py 行为一致，并修复高亮逻辑 Bug。

公开 API：
    build_dataframe(results)       构造数据 DataFrame（不含 Summary 行）
    build_summary_row(df)          构造 Summary 行 DataFrame
    write_report(...)              落盘 xlsx：列、Summary、红色高亮、饼图

修复点（PRD 明确要求）：
    原脚本 ``max_row=len(final_df) - 1`` 会跳过最后一行数据；本实现按
    ``Is Same`` 字段精确判定，跳过 Summary 行而不是固定行号。
"""
from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import pandas as pd

from .logging_setup import get_logger


COLUMN_FILENAME = "Image Filename"
COLUMN_TIMESTRING = "Time String"
COLUMN_ISSAME = "Is Same"
SUMMARY_LABEL = "Summary"


def build_dataframe(results: Dict[str, List[str]]) -> pd.DataFrame:
    """从 ``{文件名: [时间码列表]}`` 字典构造 DataFrame，按文件名排序。

    与原脚本一致：``Is Same`` 用 ``len(set(timecodes)) == 1 and bool(timecodes)``。
    """
    rows = []
    for filename in sorted(results.keys()):
        timecodes = list(results[filename])
        rows.append({
            COLUMN_FILENAME: filename,
            COLUMN_TIMESTRING: timecodes,
            COLUMN_ISSAME: len(set(timecodes)) == 1 and bool(timecodes),
        })
    return pd.DataFrame(rows, columns=[COLUMN_FILENAME, COLUMN_TIMESTRING, COLUMN_ISSAME])


def compute_summary(df: pd.DataFrame) -> Tuple[int, int, int, float]:
    """返回 (总数, true 数, false 数, 同步率)。"""
    total = len(df)
    true_count = int(df[COLUMN_ISSAME].sum()) if total > 0 else 0
    false_count = total - true_count
    sync_rate = (true_count / total) if total > 0 else 0.0
    return total, true_count, false_count, sync_rate


def build_summary_row(df: pd.DataFrame) -> pd.DataFrame:
    total, true_count, false_count, sync_rate = compute_summary(df)
    return pd.DataFrame([{
        COLUMN_FILENAME: SUMMARY_LABEL,
        COLUMN_TIMESTRING: f"False: {false_count}, True: {true_count}, Sync Rate: {sync_rate:.2%}",
        COLUMN_ISSAME: "",
    }])


def make_report_path(excel_dir: Path, prefix: str, now: datetime | None = None) -> Path:
    """生成带时间戳的 Excel 路径：``<prefix>_YYYYMMDD_HHMMSS.xlsx``。"""
    excel_dir = Path(excel_dir)
    excel_dir.mkdir(parents=True, exist_ok=True)
    ts = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return excel_dir / f"{prefix}_{ts}.xlsx"


def _draw_pie_chart(true_count: int, false_count: int, sync_rate: float, total: int) -> BytesIO:
    """画饼图并返回 PNG 字节流。"""
    import matplotlib
    matplotlib.use("Agg")  # 无 GUI 环境安全
    import matplotlib.pyplot as plt

    labels = ["Synced", "Unsynced"]
    sizes = [true_count, false_count]
    colors = ["#66B2FF", "#FF9999"]

    fig, ax = plt.subplots(figsize=(5, 4))
    ax.pie(sizes, labels=labels, autopct="%1.1f%%", startangle=90, colors=colors)
    ax.axis("equal")

    formula_text = (
        "Synchronization rate =\n"
        "True / Frames\u00d7 100%\n"
        f"= {true_count} / {total} \u00d7 100%\n"
        f"= {sync_rate:.2%}"
    )
    plt.text(1.2, 0.5, formula_text, fontsize=10, verticalalignment="center")

    buf = BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def _apply_highlight_and_chart(
    excel_path: Path,
    final_df: pd.DataFrame,
    highlight_color: str,
    true_count: int,
    false_count: int,
    sync_rate: float,
    total: int,
    insert_chart: bool,
) -> None:
    """二次打开 xlsx，按 Is Same 列加红色高亮并插入饼图。"""
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    ws = wb.active
    fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    issame_col_idx = list(final_df.columns).index(COLUMN_ISSAME) + 1  # openpyxl 1-based
    filename_col_idx = list(final_df.columns).index(COLUMN_FILENAME) + 1

    # 数据行从 Excel 第 2 行开始；最后一行是 Summary 不参与高亮判定。
    last_row = ws.max_row
    for row_idx in range(2, last_row + 1):
        filename_cell = ws.cell(row=row_idx, column=filename_col_idx)
        if filename_cell.value == SUMMARY_LABEL:
            continue
        is_same_value = ws.cell(row=row_idx, column=issame_col_idx).value
        # 判定：False 或空（非 True）的数据行高亮；True 不高亮
        if is_same_value is False or is_same_value == 0 or is_same_value == "False":
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    if insert_chart and total > 0:
        img_buf = _draw_pie_chart(true_count, false_count, sync_rate, total)
        new_col_letter = get_column_letter(ws.max_column + 1)
        ws.column_dimensions[new_col_letter].width = 25
        anchor = f"{new_col_letter}{ws.max_row + 2}"
        ws.add_image(XLImage(img_buf), anchor)

    wb.save(excel_path)


def write_report(
    results: Dict[str, List[str]],
    excel_dir: Path,
    prefix: str,
    highlight_color: str = "FFCCCC",
    *,
    now: datetime | None = None,
    insert_chart: bool = True,
) -> Path:
    """汇总写出 Excel 报告并返回最终文件绝对路径。

    Args:
        results: ``{文件名: [时间码列表]}``。
        excel_dir: 输出目录。
        prefix: 文件名前缀；最终为 ``<prefix>_YYYYMMDD_HHMMSS.xlsx``。
        highlight_color: 6 位十六进制 RGB（无 #）。
        now: 仅供测试注入固定时间戳。
        insert_chart: 是否插入饼图（测试时可关掉降低依赖）。
    """
    log = get_logger(__name__)
    df = build_dataframe(results)
    total, true_count, false_count, sync_rate = compute_summary(df)

    log.info("Frames count: %d", total)
    log.info("False count: %d True count: %d", false_count, true_count)
    log.info("Synchronization rate: %.2f%%", sync_rate * 100.0)

    summary_df = build_summary_row(df)
    final_df = pd.concat([df, summary_df], ignore_index=True)

    excel_path = make_report_path(excel_dir, prefix, now=now)
    final_df.to_excel(excel_path, index=False)

    _apply_highlight_and_chart(
        excel_path,
        final_df,
        highlight_color=highlight_color,
        true_count=true_count,
        false_count=false_count,
        sync_rate=sync_rate,
        total=total,
        insert_chart=insert_chart,
    )

    log.info(
        "Time strings and chart have been written to %s with conditional formatting and pie chart.",
        excel_path,
    )
    return excel_path
