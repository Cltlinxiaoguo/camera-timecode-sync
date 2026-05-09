import os
import cv2
from paddleocr import PaddleOCR
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import logging
from concurrent.futures import ProcessPoolExecutor, as_completed
from multiprocessing import Lock

# 新增依赖
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# 在脚本最开始记录开始时间
start_time = datetime.now()

# 初始化锁用于控制打印顺序
print_lock = Lock()

# 设置日志级别，屏蔽 PaddleOCR 的 WARNING 和 DEBUG 输出
logging.getLogger('ppocr').setLevel(logging.ERROR)

# 初始化 PaddleOCR 对象
ocr = PaddleOCR(lang='ch')


# 文件夹路径
# folder_path = r"C:\Pycahrmproject\false_drop"
folder_path = r"C:\Pycahrmproject\frames"
# folder_path = r"C:\Pycahrmproject\timecode_2026_03_04_093122"

def log_print(message):
    """带锁的日志打印函数，防止多进程输出混乱"""
    with print_lock:
        print(message)

def process_image(img_path):
    try:
        # 读取图像
        image = cv2.imread(img_path)

        # 使用整个图像进行 OCR 识别
        result = ocr.ocr(image)

        # 提取识别的文本，并筛选出符合时间格式的字符串
        time_pattern = re.compile(r'\d{2}:\d{2}:\d{2}:\d{2}')
        time_strings = [line[1][0] for line in result[0] if line[1] and time_pattern.fullmatch(line[1][0])]

        # 强制设置前两位为 '00'
        corrected_time_strings = []
        for ts in time_strings:
            parts = ts.split(':')
            if len(parts) == 4:
                parts[0] = '00'  # 强制设置 aa 为 00
                corrected_time_str = ':'.join(parts)
                corrected_time_strings.append(corrected_time_str)
            else:
                corrected_time_strings.append(ts)

        # 打印日志信息并添加分隔线
        log_print("-----------------------------------")
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # log_print(f"{current_time} {os.path.basename(img_path)}\n{corrected_time_strings}")
        log_print(f"{os.path.basename(img_path)}\n{corrected_time_strings}")
        is_same = len(set(corrected_time_strings)) == 1 and bool(corrected_time_strings)
        log_print("[✅]" if is_same else "[❌]")
        return img_path, corrected_time_strings

    except Exception as e:
        log_print("-----------------------------------")
        log_print(f"Error processing image {img_path}: {e}")
        return img_path, []

if __name__ == '__main__':
    images = []
    results_dict = {}

    for filename in os.listdir(folder_path):
        if filename.endswith('.jpg'):
            img_path = os.path.join(folder_path, filename)
            images.append(img_path)

    max_workers = min(8, os.cpu_count())  # 确保不会超过物理核心数

    print(f"🚀 开始处理 {len(images)} 张图像，使用 {max_workers} 个工作进程...")

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures_to_img = {executor.submit(process_image, img): img for img in images}
        for future in as_completed(futures_to_img):
            img_path = futures_to_img[future]
            try:
                img_path, time_strings = future.result()
                results_dict[os.path.basename(img_path)] = time_strings
            except Exception as e:
                log_print(f"Error encountered while handling future result for {img_path}: {e}")

    df_data = {
        'Image Filename': [],
        'Time String': [],
        'Is Same': []
    }

    for filename in sorted(results_dict.keys()):
        time_strings = results_dict[filename]
        df_data['Image Filename'].append(filename)
        df_data['Time String'].append(time_strings)
        df_data['Is Same'].append(len(set(time_strings)) == 1 and bool(time_strings))

    df = pd.DataFrame(df_data)

    true_count = df['Is Same'].sum()
    false_count = len(df) - true_count
    sync_rate = true_count / len(df) if len(df) > 0 else 0

    print(f"Frames count:{len(df)} ")
    print(f"False count: {false_count} True count: {true_count}")
    print(f"Synchronization rate: {sync_rate:.2%}")

    summary_df = pd.DataFrame([{
        'Image Filename': 'Summary',
        'Time String': f'False: {false_count}, True: {true_count}, Sync Rate: {sync_rate:.2%}',
        'Is Same': ''
    }])

    final_df = pd.concat([df, summary_df], ignore_index=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_filename = f'Sync_n_Timecode_{timestamp}.xlsx'
    final_df.to_excel(excel_filename, index=False)

    wb = load_workbook(excel_filename)
    ws = wb.active
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_col=3, max_row=len(final_df) - 1):
        if not row[2].value:
            for cell in row:
                cell.fill = red_fill

    # === 新增：绘制饼图并插入 Excel ===
    labels = ['Synced', 'Unsynced']
    sizes = [true_count, false_count]
    colors = ['#66B2FF', '#FF9999']

    fig, ax = plt.subplots(figsize=(5, 4))
    ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors)
    ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    # 添加公式文本
    formula_text = f"Synchronization rate =\nTrue / Frames× 100%\n= {true_count} / {len(df)} × 100%\n= {sync_rate:.2%}"
    plt.text(1.2, 0.5, formula_text, fontsize=10, verticalalignment='center')

    # 将图像转为字节流
    img_data = BytesIO()
    plt.savefig(img_data, format='png', bbox_inches='tight')
    plt.close()

    # 插入图表到 Excel 中
    img = XLImage(img_data)
    ws.column_dimensions[get_column_letter(ws.max_column + 1)].width = 25  # 自动加一列
    ws.add_image(img, f"{get_column_letter(ws.max_column + 1)}{ws.max_row + 2}")

    wb.save(excel_filename)
    print(f"Time strings and chart have been written to {excel_filename} with conditional formatting and pie chart.")

    # 程序结束时间及耗时统计
    end_time = datetime.now()
    elapsed_time = end_time - start_time  # ✅ 正确计算耗时

    print(f"程序开始时间: {start_time}")
    print(f"程序结束时间: {end_time}")
    print(f"总耗时: {elapsed_time.total_seconds():.2f} 秒")