# -*- coding: utf-8 -*-
"""Excel 写出（含 Summary、红色高亮 Bug 修复、饼图禁用）测试。

为避免依赖 matplotlib（其在 CI 中默认可用，但保险起见用 ``insert_chart=False``
覆盖核心高亮逻辑；另一个用例开启饼图以验证图片确实被插入）。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest

from camera_sync.excel_writer import (
    COLUMN_FILENAME,
    COLUMN_ISSAME,
    COLUMN_TIMESTRING,
    SUMMARY_LABEL,
    build_dataframe,
    build_summary_row,
    compute_summary,
    make_report_path,
    write_report,
)


def test_build_dataframe_sorts_and_judges():
    results = {
        "b.jpg": ["00:01:02:03", "00:01:02:03"],
        "a.jpg": ["00:01:02:03", "00:01:02:04"],
        "c.jpg": [],
    }
    df = build_dataframe(results)
    assert list(df[COLUMN_FILENAME]) == ["a.jpg", "b.jpg", "c.jpg"]
    assert list(df[COLUMN_ISSAME]) == [False, True, False]
    assert df[COLUMN_TIMESTRING].iloc[0] == ["00:01:02:03", "00:01:02:04"]


def test_compute_summary_fractions():
    df = build_dataframe({
        "a.jpg": ["x", "x"],
        "b.jpg": ["x", "y"],
        "c.jpg": ["x", "x"],
        "d.jpg": [],
    })
    total, true_c, false_c, rate = compute_summary(df)
    assert (total, true_c, false_c) == (4, 2, 2)
    assert abs(rate - 0.5) < 1e-9


def test_build_summary_row_format():
    df = build_dataframe({"a.jpg": ["x", "x"], "b.jpg": ["x", "y"]})
    sdf = build_summary_row(df)
    assert sdf.iloc[0][COLUMN_FILENAME] == SUMMARY_LABEL
    assert "Sync Rate: 50.00%" in sdf.iloc[0][COLUMN_TIMESTRING]
    assert sdf.iloc[0][COLUMN_ISSAME] == ""


def test_make_report_path_uses_timestamp(tmp_path: Path):
    fixed = datetime(2026, 5, 8, 13, 14, 15)
    p = make_report_path(tmp_path, "Sync_n_Timecode", now=fixed)
    assert p.name == "Sync_n_Timecode_20260508_131415.xlsx"
    assert p.parent == tmp_path


def test_write_report_highlights_false_rows_skips_summary(tmp_path: Path):
    """关键 PRD 修复点：跳过 Summary 行 + 按 Is Same 高亮 + 不丢最后一行数据。"""
    openpyxl = pytest.importorskip("openpyxl")

    results = {
        "a.jpg": ["00:01:02:03", "00:01:02:03"],   # True
        "b.jpg": ["00:01:02:03", "00:01:02:99"],   # False  ← 应高亮
        "c.jpg": ["00:01:02:03", "00:01:02:03"],   # True
        "z.jpg": [],                                # False  ← 最后一行数据，原脚本 bug 会漏高亮
    }
    excel_path = write_report(
        results=results,
        excel_dir=tmp_path,
        prefix="Sync_n_Timecode",
        highlight_color="FFCCCC",
        now=datetime(2026, 5, 8, 13, 14, 15),
        insert_chart=False,
    )
    assert excel_path.exists()

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    # 表头 + 4 条数据 + Summary = 6 行
    assert ws.max_row == 6

    # 找出每一行的填充情况，按文件名校验
    found = {}
    for r in range(2, ws.max_row + 1):
        fname = ws.cell(row=r, column=1).value
        fill = ws.cell(row=r, column=1).fill
        rgb = (fill.fgColor.rgb or "").upper()
        # openpyxl 保存时 rgb 通常带 alpha 前缀 'FF'，例如 'FFFFCCCC'
        is_red = rgb.endswith("FFCCCC")
        found[fname] = is_red

    assert found["a.jpg"] is False
    assert found["c.jpg"] is False
    assert found["b.jpg"] is True
    assert found["z.jpg"] is True, "原脚本 max_row=len(final_df)-1 会漏高亮最后一行，这里必须命中"
    assert found[SUMMARY_LABEL] is False, "Summary 行不应被高亮"


def test_write_report_inserts_chart(tmp_path: Path):
    """确保 insert_chart=True 时确实有图片被加入工作表。"""
    pytest.importorskip("openpyxl")
    pytest.importorskip("matplotlib")

    results = {"a.jpg": ["x", "x"], "b.jpg": ["x", "y"]}
    excel_path = write_report(
        results=results,
        excel_dir=tmp_path,
        prefix="Sync_n_Timecode",
        highlight_color="FFCCCC",
        now=datetime(2026, 5, 8, 0, 0, 0),
        insert_chart=True,
    )
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    assert len(ws._images) >= 1
